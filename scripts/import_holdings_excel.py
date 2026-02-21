"""
Import Zerodha equity and mutual fund holdings from an Excel file into the portfolio DB.
Reads two sheets: "Equity" (equity holdings + summary) and "Mutual Funds" (MF holdings + summary).
Uses each sheet's Summary: Invested Value = cost, Present Value = value.
Usage: python scripts/import_holdings_excel.py <path-to.xlsx> <date YYYY-MM-DD>
Example: python scripts/import_holdings_excel.py "h:\\Gopi\\NSE\\holdings-DG0997-Nov-02.xlsx" 2025-11-02
"""
import sys
from datetime import date
from pathlib import Path

# Run from project root so helper and config are importable
_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

import openpyxl
from helper import db


# Fallback layout (0-based column indices) when header not found
DEFAULT_HEADER_ROW = 23
COL_SYMBOL = 1
COL_QTY_AVAILABLE = 4
COL_AVG_PRICE = 9
COL_PREV_CLOSE = 10
ROW_INVESTED_VALUE = 15
ROW_PRESENT_VALUE = 16
COL_SUMMARY_LABEL = 1   # column B
COL_SUMMARY_VALUE = 2   # column C


def parse_float(val, default=0.0):
    try:
        if val is None:
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_int(val, default=0):
    try:
        if val is None:
            return default
        return int(val)
    except (TypeError, ValueError):
        return default


def _normalize_sheet_name(name):
    if name is None:
        return ""
    return str(name).strip()


def get_sheet_by_name(wb, *names):
    """Return first worksheet whose title matches one of names (case-insensitive)."""
    for want in names:
        for s in wb.worksheets:
            if _normalize_sheet_name(s.title).lower() == want.lower():
                return s
    return None


def find_header_row(ws, keyword="symbol", max_scan=35):
    """Find 1-based row number that contains keyword (e.g. 'symbol' for equity, 'scheme'/'fund' for MF)."""
    kw = keyword.strip().lower()
    for r in range(1, min(max_scan, (ws.max_row or 0) + 1)):
        row_iter = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        if not row_iter:
            continue
        full_row = row_iter[0]
        if full_row is None:
            continue
        cells = [str(v).strip().lower() if v is not None else "" for v in full_row]
        if kw in cells or any(kw in c for c in cells):
            return r
    return None


def find_summary_values(ws, max_scan=25):
    """Scan rows for 'Invested Value' and 'Present Value'; return (cost, value)."""
    cost, val = None, None
    for r in range(1, min(max_scan, (ws.max_row or 0) + 1)):
        row_iter = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        if not row_iter:
            continue
        row = row_iter[0]
        if row is None or len(row) <= COL_SUMMARY_VALUE:
            continue
        label = str(row[COL_SUMMARY_LABEL] or "").strip().lower()
        num = parse_float(row[COL_SUMMARY_VALUE])
        if "invested value" in label:
            cost = num
        elif "present value" in label:
            val = num
    return (cost, val)


def build_header_map(header_row_tuple):
    """Return dict mapping header name (lower) -> 0-based column index."""
    out = {}
    for i, cell in enumerate(header_row_tuple or []):
        name = str(cell or "").strip().lower()
        if name and name not in out:
            out[name] = i
    return out


def _get_col_idx(hmap, *candidate_names):
    for name in candidate_names:
        idx = hmap.get(name)
        if idx is not None:
            return idx
    return None


def _col(hmap, *candidate_names, fallback_idx):
    idx = _get_col_idx(hmap, *candidate_names)
    return idx if idx is not None else fallback_idx


def _merge_duplicate_holdings(holdings: list, key_fields: tuple) -> list:
    """Merge rows with same key: sum quantity, quantity-weighted avg price."""
    merged = {}
    for h in holdings:
        key = tuple(h.get(f) or "" for f in key_fields)
        if key in merged:
            m = merged[key]
            q1, q2 = m["quantity"], h["quantity"]
            total_q = q1 + q2
            if total_q <= 0:
                continue
            m["quantity"] = total_q
            m["average_price"] = round((q1 * m["average_price"] + q2 * h["average_price"]) / total_q, 4)
            m["last_price"] = h["last_price"]
        else:
            merged[key] = dict(h)
    return list(merged.values())


def load_equity_sheet(ws):
    """From Equity sheet: parse holdings and summary. Return (holdings_list, portfolio_value, portfolio_cost)."""
    header_row = find_header_row(ws, "symbol")
    if header_row is None:
        header_row = DEFAULT_HEADER_ROW
    data_start = header_row + 1
    header_tuple = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    hmap = build_header_map(header_tuple)

    idx_symbol = _col(hmap, "symbol", fallback_idx=COL_SYMBOL)
    idx_qty = _col(hmap, "quantity available", "quantity availab", "quantity av", fallback_idx=COL_QTY_AVAILABLE)
    idx_avg = _col(hmap, "average price", "average pri", fallback_idx=COL_AVG_PRICE)
    idx_prev = _col(hmap, "previous closing price", "previous closing", "previous cls", "previous close", fallback_idx=COL_PREV_CLOSE)

    cost, value = find_summary_values(ws)
    if cost is None:
        cost = parse_float(ws.cell(ROW_INVESTED_VALUE, COL_SUMMARY_VALUE + 1).value)
    if value is None:
        value = parse_float(ws.cell(ROW_PRESENT_VALUE, COL_SUMMARY_VALUE + 1).value)

    holdings = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row or 0, values_only=True):
        if not row or len(row) <= max(idx_symbol, idx_qty, idx_avg, idx_prev):
            continue
        symbol = row[idx_symbol]
        if symbol is None or (isinstance(symbol, str) and not symbol.strip()):
            continue
        symbol = str(symbol).strip()
        qty = parse_int(row[idx_qty] if idx_qty < len(row) else None)
        if qty <= 0:
            continue
        avg_price = parse_float(row[idx_avg] if idx_avg < len(row) else None)
        last_price = parse_float(row[idx_prev] if idx_prev < len(row) else None)
        if last_price <= 0:
            last_price = avg_price
        holdings.append({
            "tradingsymbol": symbol,
            "exchange": "NSE",
            "quantity": qty,
            "average_price": round(avg_price, 4),
            "last_price": round(last_price, 4),
        })

    holdings = _merge_duplicate_holdings(holdings, key_fields=("tradingsymbol", "exchange"))
    return holdings, round(value, 2), round(cost, 2)


def load_mf_sheet(ws):
    """From Mutual Funds sheet: parse holdings and summary. Return (holdings_list, mf_portfolio_value)."""
    # MF header may have: Scheme/Fund/Symbol, Folio, Units/Quantity, NAV/Price, Invested Value, Current Value
    header_row = find_header_row(ws, "scheme") or find_header_row(ws, "fund") or find_header_row(ws, "symbol")
    if header_row is None:
        header_row = DEFAULT_HEADER_ROW
    data_start = header_row + 1
    header_tuple = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    hmap = build_header_map(header_tuple)

    idx_symbol = _get_col_idx(hmap, "scheme", "fund", "fund name", "symbol", "scheme name")
    if idx_symbol is None:
        idx_symbol = COL_SYMBOL
    idx_folio = _get_col_idx(hmap, "folio", "fund folio", "folio no")
    idx_qty = _get_col_idx(hmap, "units", "quantity", "quantity available", "quantity availab")
    if idx_qty is None:
        idx_qty = COL_QTY_AVAILABLE
    idx_nav = _get_col_idx(hmap, "nav", "current nav", "price", "previous closing price", "previous close")
    if idx_nav is None:
        idx_nav = COL_PREV_CLOSE
    idx_invested = _get_col_idx(hmap, "invested value", "average price", "cost")
    if idx_invested is None:
        idx_invested = COL_AVG_PRICE

    cost, value = find_summary_values(ws)
    if value is None:
        value = parse_float(ws.cell(ROW_PRESENT_VALUE, COL_SUMMARY_VALUE + 1).value)
    mf_portfolio_value = round(value, 2) if value is not None else 0

    holdings = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row or 0, values_only=True):
        if not row or len(row) <= max(idx_symbol, idx_qty):
            continue
        symbol = row[idx_symbol]
        if symbol is None or (isinstance(symbol, str) and not symbol.strip()):
            continue
        symbol = str(symbol).strip()
        qty = parse_float(row[idx_qty] if idx_qty < len(row) else None) if idx_qty is not None else 0
        if qty <= 0:
            continue
        qty = int(qty) if qty == int(qty) else qty  # keep float if fractional units
        nav = parse_float(row[idx_nav] if idx_nav is not None and idx_nav < len(row) else None)
        avg = parse_float(row[idx_invested] if idx_invested is not None and idx_invested < len(row) else None)
        if nav <= 0:
            nav = avg
        folio = ""
        if idx_folio is not None and idx_folio < len(row) and row[idx_folio] is not None:
            folio = str(row[idx_folio]).strip()
        holdings.append({
            "tradingsymbol": symbol,
            "folio": folio or "",
            "quantity": qty,
            "average_price": round(avg, 4),
            "last_price": round(nav, 4),
        })

    holdings = _merge_duplicate_holdings(holdings, key_fields=("tradingsymbol", "folio"))
    return holdings, mf_portfolio_value


def load_holdings_from_excel(path: str):
    """Load from Equity and Mutual Funds sheets. Return (equity_list, mf_list, portfolio_value, portfolio_cost, mf_portfolio_value)."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    # read_only=False so that ws.max_row is set correctly (read_only leaves dimensions unset)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)

    equity_holdings = []
    portfolio_value = 0.0
    portfolio_cost = 0.0
    ws_equity = get_sheet_by_name(wb, "Equity")
    if ws_equity is None:
        wb.close()
        raise ValueError("Workbook has no 'Equity' sheet")
    equity_holdings, portfolio_value, portfolio_cost = load_equity_sheet(ws_equity)

    mf_holdings = []
    mf_portfolio_value = 0.0
    ws_mf = get_sheet_by_name(wb, "Mutual Funds", "Mutual Fund")
    if ws_mf is not None:
        mf_holdings, mf_portfolio_value = load_mf_sheet(ws_mf)

    wb.close()
    return equity_holdings, mf_holdings, portfolio_value, portfolio_cost, mf_portfolio_value


def main():
    if len(sys.argv) < 3:
        print("Usage: python scripts/import_holdings_excel.py <path-to.xlsx> <date YYYY-MM-DD>")
        print('Example: python scripts/import_holdings_excel.py "h:\\Gopi\\NSE\\holdings-DG0997-Nov-02.xlsx" 2025-11-02')
        sys.exit(1)
    xlsx_path = sys.argv[1]
    date_str = sys.argv[2]
    try:
        d = date.fromisoformat(date_str)
    except ValueError:
        print(f"Invalid date: {date_str}. Use YYYY-MM-DD.")
        sys.exit(1)
    equity, mf, portfolio_value, portfolio_cost, mf_portfolio_value = load_holdings_from_excel(xlsx_path)
    if not equity and not mf:
        print("No holdings found. Ensure the Excel has 'Equity' and/or 'Mutual Funds' sheets with Summary (Invested Value, Present Value) and data rows.")
        sys.exit(1)
    db.init_db()
    db.save_portfolio_day(
        d=d,
        portfolio_value=portfolio_value,
        portfolio_cost=portfolio_cost,
        buy_amount=0.0,
        sell_amount=0.0,
        month_buy=0.0,
        month_sell=0.0,
        holdings=equity,
        month_per_stock={},
        mf_holdings=mf,
        mf_portfolio_value=mf_portfolio_value,
        price_changes=None,
    )
    print(f"Imported for {date_str}: {len(equity)} equity, {len(mf)} mutual fund holdings.")
    print(f"  Equity: Present Value {portfolio_value:.2f}, Invested Value {portfolio_cost:.2f}")
    if mf:
        print(f"  Mutual Funds: Present Value {mf_portfolio_value:.2f}")


if __name__ == "__main__":
    main()
